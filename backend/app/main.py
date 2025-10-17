from datetime import datetime
import csv
import io
from typing import List

from fastapi import FastAPI, Depends, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, JSONResponse

from sqlalchemy.orm import Session

from .db import Base, engine, get_db
from . import crud, schemas
from .models import Handover

app = FastAPI(title="Skye Steward API")

# CORS for Vite dev server
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://127.0.0.1:5173", "http://localhost:5173"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("startup")
def _startup():
    # If you want safety in dev, create tables not managed by Alembic.
    Base.metadata.create_all(bind=engine)

@app.get("/health")
def health():
    return {"status": "ok"}

# ------- Handovers -------
@app.get("/api/handover", response_model=List[schemas.HandoverRead])
def get_handovers(db: Session = Depends(get_db)):
    return crud.list_handovers(db)

# ------- Guest Notes -------
@app.get("/api/guest-notes", response_model=List[schemas.GuestNoteRead])
def get_guest_notes(db: Session = Depends(get_db)):
    return crud.list_guest_notes(db)

# ------- Incidents (NEW) -------
@app.get("/api/incidents", response_model=List[schemas.IncidentRead])
def get_incidents(db: Session = Depends(get_db)):
    return crud.list_incidents(db)

@app.post("/api/incidents", response_model=schemas.IncidentRead)
def post_incident(payload: schemas.IncidentCreate, db: Session = Depends(get_db)):
    return crud.create_incident(db, payload)

# ------- Analytics -------
@app.get("/api/analytics/kpi-summary", response_model=schemas.KpiSummary)
def kpi_summary(target: float = 10000, db: Session = Depends(get_db)):
    return crud.kpi_summary(db, target)

@app.get("/api/analytics/top-items", response_model=List[schemas.TopItem])
def top_items(limit: int = 5, db: Session = Depends(get_db)):
    return crud.top_items(db, limit)

@app.get("/api/analytics/weekly")
def weekly(weeks: int = 8, db: Session = Depends(get_db)):
    return crud.weekly_revenue(db, weeks)

# ------- Exports -------
@app.get("/api/export/handovers.csv")
def export_handovers_csv(db: Session = Depends(get_db)):
    """Fix: write text CSV via StringIO (not bytes)."""
    buf = io.StringIO(newline="")
    writer = csv.writer(buf)
    writer.writerow(
        ["id", "date", "outlet", "shift", "period", "bookings", "walk_ins",
         "covers", "food_revenue", "beverage_revenue", "top_sales"]
    )
    rows: List[Handover] = crud.list_handovers(db)
    for r in rows:
        writer.writerow([
            r.id,
            r.date.isoformat(),
            r.outlet,
            r.shift,
            r.period,
            r.bookings,
            r.walk_ins,
            r.covers,
            f"{r.food_revenue:.2f}",
            f"{r.beverage_revenue:.2f}",
            ", ".join(r.top_sales or []),
        ])
    buf.seek(0)
    headers = {"Content-Disposition": "attachment; filename=handovers.csv"}
    return StreamingResponse(iter([buf.getvalue()]), media_type="text/csv", headers=headers)

@app.get("/api/export/handovers.xlsx")
def export_handovers_xlsx(db: Session = Depends(get_db)):
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Handovers"
    headers = ["id","date","outlet","shift","period","bookings","walk_ins","covers","food_revenue","beverage_revenue","top_sales"]
    ws.append(headers)

    for r in crud.list_handovers(db):
        ws.append([
            r.id, r.date.isoformat(), r.outlet, r.shift, r.period, r.bookings,
            r.walk_ins, r.covers, r.food_revenue, r.beverage_revenue, ", ".join(r.top_sales or [])
        ])
    for i in range(1, len(headers)+1):
        ws.column_dimensions[get_column_letter(i)].width = 16

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    headers_resp = {"Content-Disposition": "attachment; filename=handovers.xlsx"}
    return StreamingResponse(out, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers_resp)
